import os
import shutil
import sys
import re
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
class ExcelHandler:
    def __init__(self, template_file='LOG SHEET 3.xlsx', save_dir='logs'):
            # مسیر صحیح فایل template در حالت APK
            if getattr(sys, 'frozen', False):
                # در حالت APK اجرا می‌شود
                base_path = os.path.dirname(sys.executable)
                # یا از مسیر assets استفاده کنید
                if hasattr(sys, '_MEIPASS'):
                    base_path = sys._MEIPASS
            else:
                # در حالت عادی پایتون
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            self.template_file = os.path.join(base_path, template_file)
            
            # مسیر ذخیره در اندروید
            if os.path.exists('/storage/emulated/0/'):
                self.save_dir = '/storage/emulated/0/Documents/GT11_Logs'
            else:
                base_dir = os.path.join(os.path.expanduser('~'), 'Desktop')
                self.save_dir = os.path.join(base_dir, save_dir)
            
            os.makedirs(self.save_dir, exist_ok=True)
            
            # اطمینان از وجود فایل template
            if not os.path.exists(self.template_file):
                # اگر template نبود، یک فایل جدید ایجاد کن
                print(f"⚠️ Template not found at {self.template_file}")

    def get_save_path(self):
        """ برمی‌گردونه مسیر فایل لاگ برای تاریخ امروز """
        today = datetime.today().strftime('%Y-%m-%d')
        filename = f"GT_Log_{today}.xlsx"
        return os.path.join(self.save_dir, filename)

    def save_form_data(self, form_data):
        """ذخیره داده‌های فرم در اکسل - اگر فایل امروز وجود دارد به آن اضافه کن"""
        try:
            print("🔄 شروع فرآیند ذخیره‌سازی...")
            print(f"📋 داده‌های دریافتی: {form_data}")
            
            # بررسی وجود فایل template
            if not os.path.exists(self.template_file):
                print(f"❌ فایل template یافت نشد: {self.template_file}")
                return False
            
            # ایجاد نام فایل خروجی با تاریخ امروز
            today = datetime.now().strftime("%Y-%m-%d")
            output_path = os.path.join(self.save_dir, f"GT_Log_{today}.xlsx")
            
            # بارگذاری یا ایجاد فایل اکسل
            if os.path.exists(output_path):
                print(f"✅ فایل امروز موجود است: {output_path}")
                wb = openpyxl.load_workbook(output_path)
            else:
                print(f"📋 ایجاد فایل جدید برای امروز: {output_path}")
                shutil.copy2(self.template_file, output_path)
                wb = openpyxl.load_workbook(output_path)
            
            # انتخاب شیت دستگاه
            device_name = form_data.get('device', 'General')
            print(f"📊 دستگاه انتخاب شده: {device_name}")

            ws = self.get_worksheet(wb, device_name)
#موقت
            print("\nDEBUG: تمام نام‌های بخش‌های ممکن در شیت:")
            for r in range(1, min(80, ws.max_row + 1)):
                for c in range(1, min(6, ws.max_column + 1)):  # ستون‌های اول معمولاً نام بخش
                    val = self._get_merged_cell_value(ws, r, c)
                    if val and isinstance(val, str) and len(val.strip()) > 2:
                        print(f"  row {r:3d} | col {c:2d} ({get_column_letter(c)}) → {val.strip()}")

            
            # if device_name in wb.sheetnames:
            #     ws = wb[device_name]
            #     print(f"✅ شیت {device_name} پیدا شد")
            # else:
            #     print(f"⚠ شیت {device_name} یافت نشد، استفاده از اولین شیت")
            #     ws = wb[wb.sheetnames[0]]
            
            # پیدا کردن ستون زمان
            time_value = form_data.get('time')
            print(f"⏰ زمان انتخاب شده: {time_value}")
            
            time_col = self._find_time_column(ws, time_value)
            if time_col is None:
                print(f"❌ ستون زمان {time_value} یافت نشد")
                return False
            
            print(f"✅ ستون زمان: {time_col}")
            
            # بررسی تکراری بودن ورود داده (اختیاری - برای هشدار)
            existing_value = None
            try:
                sections_data = form_data.get('sections', {})
                if sections_data:
                    first_section = list(sections_data.keys())[0]
                    first_field = list(sections_data[first_section].keys())[0]
                    section_position = self._find_exact_section_position(ws, first_section)
                    if section_position:
                        field_column = self._find_field_column_near_section(ws, section_position)
                        if field_column:
                            field_row = self._find_exact_field_row(ws, first_field, section_position, field_column)
                            if field_row:
                                existing_value = ws.cell(row=field_row, column=time_col).value
            except Exception as e:
                print(f"⚠ خطا در بررسی تکراری بودن: {e}")
            
            if existing_value is not None:
                print(f"⚠ هشدار: این زمان قبلاً پر شده است. مقدار جدید جایگزین خواهد شد.")
            
            # پردازش و ذخیره تمام بخش‌ها و فیلدها
            sections_data = form_data.get('sections', {})
            print(f"📝 بخش‌های برای پردازش: {list(sections_data.keys())}")
            
            for section_name, fields in sections_data.items():
                print(f"\n→ تلاش برای بخش: {section_name}")
                section_position = self._find_exact_section_position(ws, section_name)
                if not section_position:
                    print(f"   ✘ بخش پیدا نشد → رد شد")
                    continue
            
                print(f"   ✓ بخش پیدا شد در row {section_position['section_row']}")
            
                field_column = self._find_field_column_near_section(ws, section_position)
                if not field_column:
                    print(f"   ✘ ستون فیلد پیدا نشد → رد شد")
                    continue
                
                for field_name, field_value in fields.items():
                    field_row = self._find_exact_field_row(ws, field_name, section_position, field_column)
                    if field_row:
                        current_value = ws.cell(row=field_row, column=time_col).value
                        if current_value is not None:
                            print(f"    ⚠ سلول ({field_row},{time_col}) قبلاً پر بود: '{current_value}' → جایگزینی")
                        # ws.cell(row=field_row, column=time_col, value=field_value)
                        self._safe_write(ws, field_row, time_col, field_value)
                        print(f"    ✅ {field_name} = {field_value} ذخیره شد")
                    else:
                        print(f"    ❌ سطر فیلد {field_name} یافت نشد")
            
            # اول: ذخیره اطلاعات شیفت (اپراتور و مهندس) در جای درست
            selected_time = form_data.get("time")
            shift_leader = form_data.get("shift_leader")      # معمولاً مهندس شیفت
            shift_engineer = form_data.get("shift_engineer")  # معمولاً اپراتور شیفت
            
            if selected_time and shift_leader and shift_engineer:
                self.write_shift_info(ws, selected_time, shift_leader, shift_engineer)
            else:
                print("⚠ اطلاعات شیفت ناقص است - ذخیره نشد")
            
            # دوم: ذخیره کامنت (در انتهای شیت، بدون تداخل با شیفت)
            comment_text = form_data.get('comment')
                    # ذخیره کامنت مربوط به شیفت (در جای درست زیر ستون شیفت)
            comment_text = form_data.get('comment')
            selected_time = form_data.get("time")

            if comment_text and selected_time:
                self.write_shift_comment(ws, selected_time, comment_text)
            elif comment_text:
                print("⚠ زمان مشخص نیست - کامنت در جای عمومی ذخیره نمی‌شود")
            
        # حذف این بخش قدیمی:
        # if comment_text:
        #     comment_row = ws.max_row + 2
        #     ws.cell(row=comment_row, column=1).value = "Comment:"
        #     ws.cell(row=comment_row, column=2).value = comment_text
            # if comment_text:
            #     comment_row = ws.max_row + 2  # دو سطر فاصله برای زیبایی
            #     ws.cell(row=comment_row, column=1).value = "Comment:"
            #     ws.cell(row=comment_row, column=2).value = comment_text
            #     print(f"📝 کامنت ذخیره شد در سطر {comment_row}")
            
            # ذخیره نهایی فایل
            wb.save(output_path)
            print(f"✅ داده‌ها با موفقیت در {output_path} ذخیره شدند")
            
            file_size = os.path.getsize(output_path) / 1024
            print(f"📊 حجم فایل: {file_size:.2f} KB")
            print(f"🕒 زمان ذخیره: {datetime.now().strftime('%H:%M:%S')}")
            
            return True
            
        except Exception as e:
            print(f"❌ خطا در ذخیره داده‌ها: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def _safe_write(self, ws, row, col, value):
        """نوشتن امن در سلول - اگر merge باشد، در سلول اصلی (بالا-چپ) می‌نویسد"""
        cell = ws.cell(row=row, column=col)

        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and
                        merged_range.min_col <= col <= merged_range.max_col):
                    target_cell = ws.cell(
                        row=merged_range.min_row,
                        column=merged_range.min_col
                    )
                    target_cell.value = value
                    print(f"   (merge) → نوشته شد در اصلی: ردیف {merged_range.min_row}, ستون {merged_range.min_col}")
                    return
            # اگر به هر دلیلی پیدا نشد
            print("   ⚠ محدوده merge پیدا نشد - تلاش مستقیم")
            cell.value = value  # این خط معمولاً نباید اجرا شود

        else:
            cell.value = value
    def get_worksheet(self, wb, device_name):
        print(f"🔍 جستجوی شیت برای دستگاه: {device_name}")

        target = self.smart_normalize(device_name)

        best_match = None
        best_score = 0

        for sheet_name in wb.sheetnames:
            norm_sheet = self.smart_normalize(sheet_name)

            target_words = set(target.split())
            sheet_words = set(norm_sheet.split())

            common = target_words & sheet_words
            score = len(common)

            if score > best_score:
                best_score = score
                best_match = sheet_name

        if best_match:
            print(f"✅ شیت match شد → {best_match}")
            return wb[best_match]

        print("⚠ هیچ شیتی match نشد → اولین شیت")
        return wb[wb.sheetnames[0]]


    def write_shift_comment(self, ws, selected_time, comment_text):
        """
    کامنت مربوط به زمان انتخاب شده را در ستون صحیح شیفت اضافه می‌کند.
    اگر قبلاً کامنتی وجود داشته باشد، کامنت جدید به انتهای آن (با خط جدید) اضافه می‌شود.
    """
        if not comment_text:
            print("⚠ کامنت خالی است - ذخیره نشد")
            return

        from openpyxl.cell.cell import MergedCell

        def norm(v):
            return str(v).strip().upper() if v else ""

        def real_cell(row, col):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                        return ws.cell(row=mr.min_row, column=mr.min_col)
            return cell

        hour = int(selected_time.split(":")[0])

        # تعیین کلید عنوان کامنت بر اساس شیفت
        if hour in (0, 4):
            comment_key = ".COMMENT"        # NIGHT SHIFT اول (چپ)
        elif hour in (8, 12, 16):
            comment_key = "COMMENT"         # DAY SHIFT (وسط)
        elif hour == 20:
            comment_key = "COMMENT."        # NIGHT SHIFT دوم (راست)
        else:
            print(f"❌ ساعت نامعتبر برای کامنت: {selected_time}")
            return

        # 1️⃣ پیدا کردن ردیف عنوان کامنت (مثل comment / .comment / comment.)
        comment_title_row = None
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if norm(ws.cell(r, c).value) == comment_key:
                    comment_title_row = r
                    break
            if comment_title_row:
                break

        if not comment_title_row:
            print(f"❌ ردیف عنوان کامنت '{comment_key}' پیدا نشد")
            return

        # 2️⃣ پیدا کردن ستون دقیق این عنوان
        comment_col = None
        for c in range(1, ws.max_column + 1):
            if norm(ws.cell(comment_title_row, c).value) == comment_key:
                comment_col = c
                break

        if not comment_col:
            print(f"❌ ستون '{comment_key}' پیدا نشد")
            return

        # 3️⃣ ردیف نوشتن کامنت = یک ردیف پایین‌تر
        value_row = comment_title_row + 1

        # 4️⃣ خواندن کامنت فعلی (اگر وجود داشته باشد)
        current_comment = real_cell(value_row, comment_col).value
        current_comment = str(current_comment) if current_comment else ""

        # 5️⃣ اضافه کردن کامنت جدید (با خط جدید اگر قبلاً چیزی بود)
        if current_comment.strip():
            # اگر قبلاً کامنت بود، یک خط جدید + زمان + کامنت جدید اضافه کن
            new_comment = f"{current_comment}\n - {comment_text.strip()}"
        else:
            # اولین کامنت
            new_comment = f" {comment_text.strip()}"

        # 6️⃣ نوشتن کامنت نهایی
        real_cell(value_row, comment_col).value = new_comment

        print(f"✅ کامنت شیفت ({selected_time}) اضافه شد به ستون {comment_col}، سطر {value_row}")
        print(f"   📝 متن نهایی: {new_comment.replace('\n', ' ↵ ')}")




    def write_shift_info(self, ws, selected_time, operator, engineer):
        from openpyxl.cell.cell import MergedCell

        def norm(v):
            return str(v).strip().upper() if v else ""

        def real_cell(row, col):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                for r in ws.merged_cells.ranges:
                    if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col:
                        return ws.cell(r.min_row, r.min_col)
            return cell

        hour = int(selected_time.split(":")[0])

        # 🔍 الگوی موردنظر بر اساس ساعت
        if hour in (0, 4):
            op_key = ".OPERATOR"
            eng_key = ".ENGINEER"
        elif hour in (8, 12, 16):
            op_key = "OPERATOR"
            eng_key = "ENGINEER"
        elif hour == 20:
            op_key = "OPERATOR."
            eng_key = "ENGINEER."
        else:
            print("❌ ساعت نامعتبر")
            return

        label_row = None

        # 1️⃣ پیدا کردن ردیفی که ENGINEER / OPERATOR داخلش هستند
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = norm(ws.cell(r, c).value)
                if v in (op_key, eng_key):
                    label_row = r
                    break
            if label_row:
                break

        if not label_row:
            print("❌ ردیف عنوان OPERATOR / ENGINEER پیدا نشد")
            return

        value_row = label_row + 1
        op_col = eng_col = None

        # 2️⃣ پیدا کردن ستون‌های دقیق
        for c in range(1, ws.max_column + 1):
            v = norm(ws.cell(label_row, c).value)
            if v == op_key:
                op_col = c
            elif v == eng_key:
                eng_col = c

        if not op_col or not eng_col:
            print("❌ ستون OPERATOR یا ENGINEER پیدا نشد")
            return

        # 3️⃣ نوشتن مقادیر (حتی اگر merge باشد)
        real_cell(value_row, op_col).value = engineer
        real_cell(value_row, eng_col).value = operator

        print(
            f"✔ {selected_time} → "
            f"OPERATOR={operator}, ENGINEER={engineer}"
        )

    # def write_shift_info(self, ws, selected_time, operator, engineer):
    #     def norm(v):
    #         return str(v).strip().upper() if v else ""

    #     def find_write_row(ws, start_row, col):
    #         r = start_row
    #         while r <= ws.max_row:
    #             v = ws.cell(r, col).value
    #             if v and "COMMENT" in str(v).upper():
    #                 return None
    #             if v in (None, ""):
    #                 return r
    #             r += 1
    #         return None

    #     hour = int(selected_time.split(":")[0])

    #     if hour in (0, 4):        # NIGHT SHIFT اول
    #         mapping = {
    #             ".OPERATOR": engineer,
    #             ".ENGINEER": operator,
    #         }
    #     elif hour in (8, 12, 16):  # DAY SHIFT
    #         mapping = {
    #             "OPERATOR": operator,
    #             "ENGINEER": engineer,
    #         }
    #     elif hour == 20:          # NIGHT SHIFT دوم
    #         mapping = {
    #             "OPERATOR.": operator,
    #             "ENGINEER.": engineer,
    #         }
    #     else:
    #         return

    #     for r in range(1, ws.max_row + 1):
    #         for c in range(1, ws.max_column + 1):
    #             key = norm(ws.cell(r, c).value)
    #             if key in mapping:
    #                 write_row = find_write_row(ws, r + 1, c)
    #                 if write_row:
    #                     ws.cell(write_row, c).value = mapping[key]

    #     print("✔ Operator / Engineer ذخیره شد | Comment کاملاً امن ماند")


   

    

    # def get_worksheet(self, wb, device_name):
    #         """پیدا کردن شیت با تطبیق انعطاف‌پذیر نام"""
    #         try:
    #             # اول نام دقیق رو چک کن
    #             if device_name in wb.sheetnames:
    #                 return wb[device_name]
                
    #             # اگر دقیق پیدا نشد، به صورت case-insensitive جستجو کن
    #             device_lower = device_name.lower()
    #             for sheet_name in wb.sheetnames:
    #                 if sheet_name.lower() == device_lower:
    #                     return wb[sheet_name]
                
    #             # اگر هنوز پیدا نشد، جستجوی partial انجام بده
    #             for sheet_name in wb.sheetnames:
    #                 if device_lower in sheet_name.lower():
    #                     print(f"✅ شیت مشابه پیدا شد: '{sheet_name}' برای دستگاه '{device_name}'")
    #                     return wb[sheet_name]
                
    #             # همه شیت‌ها رو برای دیباگ نمایش بده
    #             print(f"📋 تمام شیت‌های موجود: {wb.sheetnames}")
    #             return wb[wb.sheetnames[0]]
                
    #         except Exception as e:
    #             print(f"❌ خطا در پیدا کردن شیت: {e}")
    #             return wb[wb.sheetnames[0]]


    # def _get_merged_cell_value(self, worksheet, row, col):
    #     for mr in worksheet.merged_cells.ranges:
    #         if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
    #             return worksheet.cell(row=mr.min_row, column=mr.min_col).value
    #     return worksheet.cell(row=row, column=col).value

    
    def _get_merged_cell_value(self, worksheet, row, col):
        from openpyxl.cell.cell import MergedCell

        cell = worksheet.cell(row=row, column=col)

        if isinstance(cell, MergedCell):
            for mr in worksheet.merged_cells.ranges:
                if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                    return worksheet.cell(row=mr.min_row, column=mr.min_col).value

        return cell.value


    def smart_normalize(self, text):
    
        if not text:
            return ""
        
        text = str(text).lower()

        text = text.replace("&", " and ")
        text = text.replace("\n", " ").replace("\r", " ")

        text = re.sub(r"[^\w\s]", " ", text)
        text = " ".join(text.split())

        return text


    def _text_to_words(self, text):
        norm = self._smart_normalize(text)
        return set(norm.split())

    def _smart_match(self, cell_text, field_text):
        cell_words = self._text_to_words(cell_text)
        field_words = self._text_to_words(field_text)

        if not cell_words or not field_words:
            return False

        # اگر بیشتر کلمات مشترک باشند → match
        common = cell_words & field_words
        similarity = len(common) / len(field_words)

        return similarity >= 0.6   # آستانه تطبیق (60%)

    # def _find_exact_field_row(self, worksheet, field_name, section_position, field_column):
    #     """
    #     جستجو برای نام فیلد در ستون field_column؛
    #     شروع از همان ردیف section_position['section_row'] (در صورت قرار گرفتن فیلد روی همان ردیف)
    #     و استفاده از _get_merged_cell_value برای خواندن صحیح مقادیر merge شده.
    #     """
    #     try:
    #         section_row = section_position['section_row']
    #         mapped_name = self.map_field_name(field_name)
    #         search_str = str(mapped_name).lower()

    #         print(f"🔍 پیدا کردن فیلد '{field_name}' near بخش در سطر {section_row}")

    #         # از خود سطر بخش شروع کن (نه سطر بعدی)
    #         start_row = section_row
    #         end_row = min(section_row + 20, worksheet.max_row)  # محدودهٔ جستجو تا ۲۰ سطر پایین‌تر

    #         for row in range(start_row, end_row + 1):
    #             cell_value = self._get_merged_cell_value(worksheet, row, field_column)

    #             if cell_value:
    #                 cell_str = str(cell_value).lower().replace(" ", "")
    #                 search_clean = search_str.lower().replace(" ", "")

    #                 if search_clean in cell_str:
    #                     print(f"✅ فیلد '{mapped_name}' در سطر {row} پیدا شد: '{cell_value}'")
    #                     return row


    #         print(f"❌ فیلد '{field_name}' در محدوده بخش یافت نشد")
    #         return None

    #     except Exception as e:
    #         print(f"❌ خطا در پیدا کردن فیلد: {e}")
    #         return None


    def _find_exact_field_row(self, worksheet, field_name, section_position, field_column):
        try:
            section_row = section_position['section_row']
            mapped_name = self.map_field_name(field_name)

            target = self.smart_normalize(mapped_name)

            print(f"🔍 جستجوی هوشمند فیلد '{field_name}' از سطر {section_row}")

            start_row = section_row
            end_row = min(section_row + 30, worksheet.max_row)

            best_match = None
            best_score = 0

            for row in range(start_row, end_row + 1):
                cell_value = self._get_merged_cell_value(worksheet, row, field_column)

                if not cell_value:
                    continue

                cell_norm = self.smart_normalize(cell_value)

                target_words = set(target.split())
                cell_words = set(cell_norm.split())

                common = target_words & cell_words
                score = len(common)

                if score > best_score:
                    best_score = score
                    best_match = row

            if best_match and best_score >= 1:
                print(f"✅ فیلد پیدا شد → سطر {best_match}")
                return best_match

            print(f"❌ فیلد '{field_name}' پیدا نشد")
            return None

        except Exception as e:
            print(f"❌ خطا در پیدا کردن فیلد: {e}")
            return None


    # def _find_exact_section_position(self, worksheet, section_name):
    #     """پیدا کردن موقعیت دقیق یک بخش در اکسل"""
    #     try:
    #         print(f"🔍 پیدا کردن موقعیت دقیق بخش: {section_name}")
            
    #         # ابتدا به دنبال نام دقیق بخش بگرد
    #         for row in range(1, worksheet.max_row + 1):
    #             for col in range(1, worksheet.max_column + 1):
    #                 cell_value = worksheet.cell(row=row, column=col).value
                    
    #                 if cell_value and str(cell_value).strip().lower() == section_name.lower():
    #                     print(f"✅ بخش {section_name} در سطر {row}, ستون {col} پیدا شد")
    #                     return {'section_row': row, 'section_col': col}
            
    #         # اگر نام دقیق پیدا نشد، به دنبال بخشی بگرد که شامل این نام باشد
    #         for row in range(1, worksheet.max_row + 1):
    #             for col in range(1, worksheet.max_column + 1):
    #                 cell_value = worksheet.cell(row=row, column=col).value
                    
    #                 if cell_value and section_name.lower() in str(cell_value).lower():
    #                     print(f"✅ بخش مشابه {section_name} در سطر {row}, ستون {col} پیدا شد: '{cell_value}'")
    #                     return {'section_row': row, 'section_col': col}
            
    #         print(f"❌ بخش {section_name} یافت نشد")
    #         return None
            
    #     except Exception as e:
    #         print(f"❌ خطا در پیدا کردن موقعیت بخش: {e}")
    #         return None


    #نسخه اصلی !!!!!!!!!!!!!!!!!!!!!!
    # def _find_exact_section_position(self, worksheet, section_name):
    #     print(f"🔍 جستجوی هوشمند بخش: {section_name}")

    #     target = self.smart_normalize(section_name)

    #     best_match = None
    #     best_score = 0

    #     for row in range(1, worksheet.max_row + 1):
    #         for col in range(1, worksheet.max_column + 1):
    #             cell_value = self._get_merged_cell_value(worksheet, row, col)

    #             if not cell_value:
    #                 continue

    #             cell_norm = self.smart_normalize(cell_value)

    #             target_words = set(target.split())
    #             cell_words = set(cell_norm.split())

    #             common = target_words & cell_words
    #             score = len(common)

    #             if score > best_score:
    #                 best_score = score
    #                 best_match = (row, col, cell_value)

    #     if best_match and best_score >= 2:
    #         row, col, txt = best_match
    #         print(f"✅ بخش پیدا شد → سطر {row}, ستون {col} : '{txt}'")
    #         return {'section_row': row, 'section_col': col}

    #     print(f"❌ بخش '{section_name}' پیدا نشد")
    #     return None


    #موقت
    def _find_exact_section_position(self, worksheet, section_name):
        print(f"🔍 جستجوی بخش: '{section_name}'")

        target = self.smart_normalize(section_name)

        best_match = None
        best_score = 0

        for row in range(1, worksheet.max_row + 1):
            for col in range(1, min(10, worksheet.max_column + 1)):  # جستجو در ستون‌های بیشتر
                cell_value = self._get_merged_cell_value(worksheet, row, col)
                if not cell_value:
                    continue

                cell_norm = self.smart_normalize(cell_value)

                target_words = set(target.split())
                cell_words = set(cell_norm.split())

                common = target_words & cell_words
                score = len(common)

                if score > best_score:
                    best_score = score
                    best_match = (row, col, cell_value)

        # آستانه را خیلی پایین آوردیم تا تست راحت‌تر باشد
        if best_match and best_score >= 1:   # ← اینجا 1 یا حتی 0.5 تست کن
            row, col, txt = best_match
            print(f"✅ پیدا شد → row {row}, col {col} : '{txt}' (score={best_score})")
            return {'section_row': row, 'section_col': col}

        print(f"❌ بخش پیدا نشد (بهترین score = {best_score})")
        return None
    def _find_field_column_near_section(self, worksheet, section_position):
        """پیدا کردن ستون فیلدها در نزدیکی بخش"""
        try:
            section_row = section_position['section_row']
            section_col = section_position['section_col']
            
            print(f"🔍 پیدا کردن ستون فیلدها near بخش در سطر {section_row}")
            
            # معمولاً فیلدها در ستون‌های سمت راست بخش هستند
            # ابتدا ستون‌های مجاور را بررسی کن
            for col_offset in [1, 2, 3, -1, -2, -3]:  # راست و چپ را بررسی کن
                check_col = section_col + col_offset
                if 1 <= check_col <= worksheet.max_column:
                    cell_value = worksheet.cell(row=section_row, column=check_col).value
                    if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 2:
                        print(f"✅ ستون فیلدها پیدا شد: {check_col} ('{cell_value}')")
                        return check_col
            
            # اگر در ردیف بخش پیدا نشد، در ردیف‌های پایین‌تر جستجو کن
            for row in range(section_row + 1, min(section_row + 10, worksheet.max_row + 1)):
                for col in range(1, min(10, worksheet.max_column + 1)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 2:
                        # چک کن که این سلول مربوط به واحدها نباشد
                        unit_indicators = ['ºc', 'bar', 'v', 'a', 'kv', 'ka', 'mw', 'mbar', '%', 'ok/not ok', 'auto/manual']
                        if not any(indicator in str(cell_value).lower() for indicator in unit_indicators):
                            print(f"✅ ستون فیلدها پیدا شد: {col} در سطر {row} ('{cell_value}')")
                            return col
            
            print("❌ ستون فیلدها یافت نشد")
            return None
            
        except Exception as e:
            print(f"❌ خطا در پیدا کردن ستون فیلدها: {e}")
            return None

    def debug_all_columns(self, worksheet, max_rows=20, max_cols=10):
        """دیباگ تمام ستون‌های مهم"""
        print("🔍 دیباگ تمام ستون‌های مهم:")
        
        for col in range(1, min(max_cols + 1, worksheet.max_column + 1)):
            print(f"\n📊 ستون {col}:")
            has_data = False
            
            for row in range(1, min(max_rows + 1, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    print(f"   سطر {row}: '{cell_value}'")
                    has_data = True
            
            if not has_data:
                print("   (بدون داده)")

    def map_field_name(self, field_name):
        """تطبیق نام فیلد با نام‌های موجود در اکسل"""
        field_mapping = {
            'battery voltage': 'Battery voltage',
            'coolant temp': 'coolant temp', 
            'level of fuel': 'Level of fuel',
            'leakage': 'Leakage',
            'dg model(local panel)': 'DG mode(Local panel)',
            'dg mode(local panel)': 'DG mode(Local panel)',
            'dg mode(8610)': 'DG mode(8610)',
            
            # برای سایر دستگاه‌ها
            'voltage': 'Voltage',
            'current': 'Current', 
            'temperature': 'Temperature',
            'pressure': 'Pressure',
            'level': 'Level',
            'status': 'Status'
        }
        
        lower_field = field_name.lower()
        return field_mapping.get(lower_field, field_name)

    def _find_field_row(self, worksheet, field_name):
        """پیدا کردن سطر مربوط به یک فیلد خاص"""
        try:
            print(f"🔍 جستجوی فیلد: '{field_name}'")
            
            # جستجو در تمام سطرها و ستون‌ها
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    
                    if cell_value and str(field_name).lower() in str(cell_value).lower():
                        print(f"✅ فیلد '{field_name}' در سطر {row}, ستون {col} پیدا شد: '{cell_value}'")
                        return row
            
            print(f"❌ فیلد '{field_name}' در هیچ‌یک از سلول‌ها یافت نشد")
            return None
            
        except Exception as e:
            print(f"❌ خطا در پیدا کردن سطر فیلد: {e}")
            return None




    # def get_shift_by_time(time_str):
    #     hour = int(time_str.split(":")[0])

    #     if hour in [0, 4]:
    #         return "night_left"
    #     elif hour in [8, 12, 16]:
    #         return "day"
    #     elif hour == 20:
    #         return "night_right"
    #     else:
    #         raise ValueError("Unknown time! Should be 0,4,8,12,16,20")

    # def find_shift_columns(sheet):
    #     shift_row = None
    #     shift_cols = {
    #         "night_left": None,
    #         "day": None,
    #         "night_right": None
    #     }

    #     # 1) پیدا کردن ردیف شیفت‌ها
    #     for row in sheet.iter_rows(min_row=1, max_row=50):
    #         for cell in row:
    #             if cell.value in ["NIGHT SHIFT", "DAY SHIFT"]:
    #                 shift_row = cell.row
    #                 break
    #         if shift_row:
    #             break

    #     if not shift_row:
    #         raise ValueError("Shift row not found!")

    #     # 2) پیدا کردن ستون‌های سه شیفت
    #     for cell in sheet[shift_row]:
    #         if cell.value == "NIGHT SHIFT":
    #             if shift_cols["night_left"] is None:
    #                 shift_cols["night_left"] = cell.column
    #             else:
    #                 shift_cols["night_right"] = cell.column

    #         elif cell.value == "DAY SHIFT":
    #             shift_cols["day"] = cell.column

    #     # 3) ردیف اپراتور/مهندس = یک ردیف پایین‌تر
    #     operator_row = shift_row + 1

    #     return shift_cols, shift_row, operator_row


    # def write_value(sheet, description_cell, value, time_str):
    #     shift_cols, shift_row, operator_row =find_shift_columns(sheet)
    #     shift_key = get_shift_by_time(time_str)
    #     target_col = shift_cols[shift_key]

    # # same row as description
    #     row = description_cell.row

    # # نوشتن مقدار در سلول صحیح
    #     sheet.cell(row=row, column=target_col).value = value


    def _resolve_merged_cell(ws, cell):
        """اگر cell یک MergedCell هست، سلول بالا-چپِ محدودهٔ merge رو برگردون."""
        if isinstance(cell, MergedCell):
            for mr in ws.merged_cells.ranges:
                if mr.min_row <= cell.row <= mr.max_row and mr.min_col <= cell.column <= mr.max_col:
                    return ws.cell(row=mr.min_row, column=mr.min_col)
        return cell

    def find_shift_columns(ws):
        """
        دنبال ردیفِ header می‌گرده (شامل 'NIGHT SHIFT' یا 'DAY SHIFT') و
        ستون‌های OPERATOR و ENGINEER رو برای left/day/right برمی‌گردونه.
        خروجی: (shift_cols, header_row, operator_header_row)
        shift_cols = {
        'night_left': {'operator': col_index, 'engineer': col_index},
        'day': {'operator': col_index, 'engineer': col_index},
        'night_right': {'operator': col_index, 'engineer': col_index},
        }
        """
        shift_cols = {
            'night_left': {'operator': None, 'engineer': None},
            'day': {'operator': None, 'engineer': None},
            'night_right': {'operator': None, 'engineer': None},
        }
        header_row = None

        # ابتدا ردیفی که شامل NIGHT SHIFT / DAY SHIFT هست رو پیدا کن
        for r in range(1, ws.max_row + 1):
            row_text = " ".join(
                str(ws.cell(r, c).value or "").strip().lower() for c in range(1, min(20, ws.max_column) + 1)
            )
            if 'night shift' in row_text or 'day shift' in row_text:
                header_row = r
                break

        if not header_row:
            # fallback: اگر پیدا نشد، سعی کن ردیف پایین‌تر یا بالا رو چک کنی (کاربر گفت ممکنه متفاوت باشه)
            header_row = max(1, ws.max_row - 20)

        # در ردیف بعدی معمولاً سلول‌های OPERATOR / ENGINEER مرج شده قرار دارن.
        # پس چند ردیف پایین‌تر را اسکن می‌کنیم تا عناوین OPERATOR/ENGINEER را پیدا کنیم.
        search_start = header_row + 1
        search_end = min(ws.max_row, header_row + 6)

        # ذخیرهٔ candidate ها به ترتیب از چپ به راست
        candidates = []
        for c in range(1, ws.max_column + 1):
            for r in range(search_start, search_end + 1):
                val = ws.cell(r, c).value
                if val and isinstance(val, str):
                    txt = val.strip().lower()
                    if txt in ('operator', 'engineer'):
                        candidates.append((c, r, txt))
        # حالا بر اساس موقعیت candidates را به سه گروه تقسیم میکنیم (left, mid, right)
        if candidates:
            # گروه‌بندی بر اساس ستون
            cols = sorted(set(c for c, r, t in candidates))
            if not cols:
                return shift_cols, header_row, search_start

            # تقسیم سه ناحیه: left (اولی)، mid (وسط)، right (آخر)
            # اگر کمتر از 3 ستون باشه، سعی می‌کنیم مطابق ترتیب قرار بدیم
            unique_cols = cols
            if len(unique_cols) == 1:
                # همهٔ OPERATOR/ENGINEER فقط در یک ناحیه است — قرار میدیم در day
                left_col = mid_col = right_col = unique_cols[0]
            elif len(unique_cols) == 2:
                left_col = unique_cols[0]
                mid_col = unique_cols[1]
                right_col = unique_cols[1]
            else:
                left_col = unique_cols[0]
                mid_col = unique_cols[len(unique_cols)//2]
                right_col = unique_cols[-1]

            # حالا به ازای هر candidate قرارشون در منطقه مناسب بگذار
            for c, r, txt in candidates:
                # تعیین ناحیه
                if c <= (left_col + mid_col) // 2:
                    region = 'night_left'
                elif c >= (mid_col + right_col) // 2:
                    region = 'night_right'
                else:
                    region = 'day'

                if txt == 'operator':
                    shift_cols[region]['operator'] = c
                elif txt == 'engineer':
                    shift_cols[region]['engineer'] = c

        return shift_cols, header_row, search_start

    def get_shift_by_time(time_str):
        # normalize
        if ":" in time_str:
            h = time_str.split(":")[0]
        else:
            h = str(time_str)
        h = h.lstrip("0") or "0"
        try:
            hh = int(h)
        except:
            return None

        if hh in (0, 4):
            return 'night_left'
        if hh in (8, 12, 16):
            return 'day'
        if hh == 20:
            return 'night_right'
        return None

    def write_value(self, ws, description_cell_coord, role, value, time_str):
        """
        ws: worksheet
        description_cell_coord: نمونه "A17" — از این سلول ردیف هدف گرفته میشه
        role: 'operator' یا 'engineer'
        value: رشته‌ای که باید نوشته بشه
        time_str: مثلاً "08:00" یا "8:00" یا "20:00"
        """

        # نرمال‌سازی زمان -> فقط ساعت بصورت عددی (بدون صفر پیشرو)
        if time_str is None:
            print("❌ write_value: time_str is None")
            return

        if ":" in str(time_str):
            hour = str(time_str).split(":")[0]
        else:
            hour = str(time_str)

        hour = hour.lstrip("0") or "0"   # "08" -> "8", "00" -> "0"

        # تعیین ستون‌های هدف بر پایه ساعت
        # هر ورودی mapping -> (operator_col, engineer_col)
        shift_map = {
            ('0', '4'): ('B', 'C'),      # NIGHT LEFT
            ('8', '12', '16'): ('E', 'F'), # DAY SHIFT
            ('20',): ('H', 'I')         # NIGHT RIGHT
        }

        target_cols = None
        for keys, cols in shift_map.items():
            if hour in keys:
                target_cols = cols
                break

        if target_cols is None:
            print("❌ write_value: ساعت نامعتبر:", time_str)
            return

        # انتخاب ستون براساس نقش
        if role == 'operator':
            target_col_letter = target_cols[0]
        elif role == 'engineer':
            target_col_letter = target_cols[1]
        else:
            print("❌ write_value: role نامعتبر:", role)
            return

        # پیدا کردن ردیف از description_cell_coord
        try:
            description_cell = ws[description_cell_coord]
            target_row = description_cell.row
        except Exception:
            # اگر coord اشتباه باشه سعی کن عدد انتهای رشته رو استخراج کنی
            import re
            m = re.search(r'(\d+)$', str(description_cell_coord))
            if m:
                target_row = int(m.group(1))
            else:
                target_row = ws.max_row

        target_coord = f"{target_col_letter}{target_row}"
        # اگر سلول target، یک MergedCell باشه باید به سلول بالا-چپ merge رفته و مقدار رو اونجا بنویسیم
        from openpyxl.cell.cell import MergedCell
        cell_obj = ws[target_coord]
        if isinstance(cell_obj, MergedCell):
            # پیدا کردن محدوده merge و گرفتن سلول بالا-چپ
            for mr in ws.merged_cells.ranges:
                if mr.min_row <= cell_obj.row <= mr.max_row and mr.min_col <= cell_obj.column <= mr.max_col:
                    real = ws.cell(row=mr.min_row, column=mr.min_col)
                    real.value = value
                    # اگر لازم بود می‌تونیم style هم منتقل کنیم
                    return
            # اگر نیافتیم fallback:
            ws[target_coord].value = value
        else:
            ws[target_coord].value = value



    def debug_time_row(self, worksheet):
        """برای دیباگ کردن سطر اول (زمان‌ها)"""
        print("🔍 دیباگ سطر اول (زمان‌ها):")
        
        time_row = 1
        row_data = []
        
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=time_row, column=col).value
            if cell_value is not None:
                row_data.append(f"ستون {col}: '{cell_value}'")
        
        print(f"سطر 1: {', '.join(row_data)}")
        
        # نمایش کامل‌تر برای دیباگ
        print("\n📋 تمام سلول‌های سطر اول:")
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=time_row, column=col).value
            print(f"ستون {col} ({get_column_letter(col)}): '{cell_value}'")

    # پیدا کردن سطر واقعی زمان
    def _find_time_row(self, ws):
        for row in range(1, 10):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if val and ("0:00" in str(val) or "4:00" in str(val) or "8:00" in str(val)):
                    return row
        return 1  # fallback

    def _find_time_column(self, ws, time_value):
        time_num = int(time_value.split(":")[0])
        time_row = self._find_time_row(ws)

        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=time_row, column=col).value
            if cell_value and str(time_num) in str(cell_value):
                return col

        return None

    def find_field_column(self, worksheet):
        """پیدا کردن ستونی که شامل فیلدها است"""
        try:
            print("🔍 در حال پیدا کردن ستون فیلدها...")
            
            # بررسی 20 سطر اول و 10 ستون اول
            for col in range(1, min(11, worksheet.max_column + 1)):
                field_count = 0
                print(f"📊 بررسی ستون {col}:")
                
                for row in range(1, min(21, worksheet.max_row + 1)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 3:
                        print(f"   سطر {row}: '{cell_value}'")
                        field_count += 1
                
                # اگر در این ستون چندین فیلد پیدا شد، احتمالاً ستون فیلدها است
                if field_count >= 3:  # حداقل 3 فیلد پیدا شده
                    print(f"✅ ستون فیلدها پیدا شد: ستون {col}")
                    return col
            
            print("❌ ستون فیلدها یافت نشد")
            return None
            
        except Exception as e:
            print(f"❌ خطا در پیدا کردن ستون فیلدها: {e}")
            return None

    def _find_field_row(self, worksheet, field_name, field_column=None):
        """پیدا کردن سطر مربوط به یک فیلد خاص"""
        try:
            # اگر ستون فیلدها مشخص نیست، آن را پیدا کن
            if field_column is None:
                field_column = self.find_field_column(worksheet)
                if field_column is None:
                    print("❌ ستون فیلدها یافت نشد")
                    return None
            
            print(f"🔍 جستجوی فیلد: '{field_name}' در ستون {field_column}")
            
            # تطبیق نام فیلد
            mapped_name = self.map_field_name(field_name)
            search_str = str(mapped_name).lower()
            
            # در ستون فیلدها جستجو کن
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=field_column).value
                
                if cell_value:
                    cell_str = str(cell_value).lower()
                    
                    if search_str in cell_str:
                        print(f"✅ فیلد '{mapped_name}' در سطر {row} پیدا شد: '{cell_value}'")
                        return row, field_column
            
            print(f"❌ فیلد '{mapped_name}' در ستون {field_column} یافت نشد")
            return None, field_column
            
        except Exception as e:
            print(f"❌ خطا در پیدا کردن سطر فیلد: {e}")
            return None, field_column

    def _find_sections_for_device(self, worksheet, device_row, device_col):
        """پیدا کردن تمام بخش‌های یک دستگاه"""
        try:
            print(f"🔍 پیدا کردن بخش‌های دستگاه در سطر {device_row}")
            
            sections = []
            current_section = None
            
            # از سطر دستگاه به بعد رو بررسی کن
            for row in range(device_row + 1, worksheet.max_row + 1):
                # سلول همستون با دستگاه رو بررسی کن (معمولاً بخش‌ها در همین ستون هستند)
                section_cell = worksheet.cell(row=row, column=device_col).value
                
                if section_cell and section_cell.strip():
                    # اگر سلول خالی نیست، ممکنه یک بخش جدید باشه
                    if not current_section or section_cell != current_section['name']:
                        current_section = {
                            'name': section_cell,
                            'start_row': row,
                            'end_row': None,
                            'field_column': None
                        }
                        sections.append(current_section)
                        print(f"✅ بخش جدید پیدا شد: {section_cell} در سطر {row}")
                
                # اگر بخش جاری داریم، ستون فیلدهای اون رو پیدا کن
                if current_section and current_section['field_column'] is None:
                    field_col = self._find_field_column_for_row(worksheet, row)
                    if field_col:
                        current_section['field_column'] = field_col
                        print(f"✅ ستون فیلدها برای بخش {current_section['name']}: {field_col}")
            
            return sections
            
        except Exception as e:
            print(f"❌ خطا در پیدا کردن بخش‌ها: {e}")
            return []

    def _find_device_position(self, worksheet, device_name):
        """پیدا کردن موقعیت دستگاه در شیت"""
        try:
            print(f"🔍 پیدا کردن موقعیت دستگاه: {device_name}")
            
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    
                    if cell_value and device_name.lower() in str(cell_value).lower():
                        print(f"✅ دستگاه {device_name} در سطر {row}, ستون {col} پیدا شد: '{cell_value}'")
                        return row, col
            
            print(f"❌ دستگاه {device_name} یافت نشد")
            return None, None
            
        except Exception as e:
            print(f"❌ خطا در پیدا کردن دستگاه: {e}")
            return None, None

    def find_device_structure(self, worksheet, device_name):
        """پیدا کردن ساختار سلسله مراتبی دستگاه و بخش‌های آن"""
        try:
            print(f"🔍 پیدا کردن ساختار دستگاه: {device_name}")
            
            device_row, device_col = self._find_device_position(worksheet, device_name)
            if not device_row:
                print(f"❌ دستگاه {device_name} یافت نشد")
                return None
            
            print(f"✅ دستگاه {device_name} در سطر {device_row}, ستون {device_col} پیدا شد")
            
            # پیدا کردن تمام بخش‌های این دستگاه
            sections = self._find_sections_for_device(worksheet, device_row, device_col)
            
            return {
                'device_row': device_row,
                'device_col': device_col,
                'sections': sections
            }
            
        except Exception as e:
            print(f"❌ خطا در پیدا کردن ساختار دستگاه: {e}")
            return None

    def _save_operator_info(self, worksheet, form_data):
        """ذخیره اطلاعات اپراتورها در اکسل"""
        try:
            operator_info = {
                'shift': form_data.get('shift', ''),
                'shift_leader': form_data.get('shift_leader', ''),
                'shift_engineer': form_data.get('shift_engineer', ''),
                'date': form_data.get('date', '')
            }
            
            # اطلاعات معمولاً در سطرهای 1-5 قرار می‌گیرند
            for row in range(1, 6):
                cell_value = worksheet.cell(row=row, column=1).value
                
                if cell_value:
                    cell_str = str(cell_value).lower()
                    
                    if 'shift' in cell_str and 'engineer' not in cell_str:
                        worksheet.cell(row=row, column=2).value = operator_info['shift']
                    elif 'leader' in cell_str:
                        worksheet.cell(row=row, column=2).value = operator_info['shift_leader']
                    elif 'engineer' in cell_str:
                        worksheet.cell(row=row, column=2).value = operator_info['shift_engineer']
                    elif 'date' in cell_str:
                        worksheet.cell(row=row, column=2).value = operator_info['date']
                        
            print("✅ اطلاعات اپراتورها ذخیره شد")
                
        except Exception as e:
            print(f"⚠ خطا در ذخیره اطلاعات اپراتورها: {e}")

    def _check_field_status(self, device, section_name, field_name, field_value):
        """بررسی اینکه مقدار فیلد در محدوده مجاز است یا نه"""
        try:
            # این بخش باید با توجه به ساختار device_sections_map شما تنظیم شود
            # فعلاً یک نمونه ساده پیاده‌سازی می‌کنم
            
            # اگر مقدار خالی است، OK در نظر بگیر
            if field_value is None or field_value == "":
                return 'OK'
                
            # برای مقادیر عددی، بررسی محدوده
            try:
                numeric_value = float(field_value)
                # اینجا می‌توانید محدوده‌های خاص هر فیلد را بررسی کنید
                # به عنوان مثال ساده:
                if numeric_value < 0:
                    return 'ERROR'
            except ValueError:
                # اگر مقدار عددی نبود، OK در نظر بگیر
                pass
                
            return 'OK'
            
        except Exception as e:
            print(f"⚠ خطا در بررسی وضعیت فیلد: {e}")
            return 'OK'

